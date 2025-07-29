# -*- coding: utf-8 -*-
import os
import math
import pandas as pd
from openpyxl import Workbook
from collections import defaultdict
import sys

# 打开日志文件
log_file = open(r'/home/shanli/video_analysis/behavior_result/male_20230825/output.log', 'a')
sys.stdout = log_file


def txt(folder_path, target_folder, save_folder):
    try:
        # target_contents = []
        # all_coordinates = []

        for filename in os.listdir(folder_path):
            if filename.endswith('.txt'):
                file_path = os.path.join(folder_path, filename)
                target_file_path = os.path.join(target_folder, filename)
                print(target_file_path)

                target_contents = []
                if os.path.exists(target_file_path):
                    with open(target_file_path, 'r', encoding='utf-8') as target_file:
                        for line in target_file:
                            line = line.strip()
                            if line:
                                x, y = map(float, line.split(','))
                                target_contents.append((x, y))

                all_coordinates = []
                with open(file_path, 'r', encoding='utf-8') as file:
                    for line in file:
                        line = line.strip()
                        if line:
                            numbers = list(map(int, line.split(',')))
                            all_coordinates.append(numbers)

                os.makedirs(save_folder, exist_ok=True)
                filename = os.path.splitext(filename)[0]
                save_path = os.path.join(save_folder, f"{filename}.xlsx")
                wb = Workbook()
                ws = wb.active

                ws.append([
                    'Mouse Name', 'Eating Duration', 'Drinking Duration', 'Climbing Duration',
                    'Curling Duration', 'Stretching Duration', 'Bending Duration',
                    'Grooming Duration', 'Standing Duration', 'Still Duration',
                    'Walking Duration', 'Walking Distance', 'Rest Time', 'Rest Count',
                    'Sleep Time', 'Sleep Count', 'mouse12', 'mouse13', 'mouse23',
                    'Group of three mice'
                ])

                # 添加 mouse 标签到第二、第三和第四行
                ws.append(['mouse1', '', ''])  # 第二行
                ws.append(['mouse2', '', ''])  # 第三行
                ws.append(['mouse3', '', ''])  # 第四行

                # 分组数据
                all_coordinates_3m = []
                all_coordinates_2m = []
                all_coordinates_1m = []

                frame_groups = defaultdict(list)

                # 将数据按帧号分组
                for data in all_coordinates:
                    frame_number = data[0]  # 假设 data[0] 是帧号
                    frame_groups[frame_number].append(data)

                # 将帧数据分为三只、两只和一只老鼠
                for frame_number, frame_data in frame_groups.items():
                    # 通过老鼠的 ID 来判断有多少只老鼠
                    num_mice = len(set(data[1] for data in frame_data))  # 假设 data[1] 是老鼠的 ID

                    # 根据老鼠的数量，分配到对应的列表
                    if num_mice == 3:
                        all_coordinates_3m.extend(frame_data)
                    elif num_mice == 2:
                        all_coordinates_2m.extend(frame_data)
                    elif num_mice == 1:
                        all_coordinates_1m.extend(frame_data)
                print(len(all_coordinates_3m), len(all_coordinates_2m), len(all_coordinates_1m))

                # 已经定义了 behavior、behavior2mous 和 behavior1mous 函数

                # 初始化统计变量
                total_mouse1 = []
                total_mouse2 = []
                total_mouse3 = []

                if len(all_coordinates_3m) == 0:
                    sleep_time1, sleep_count1 = 0,0
                    rest_time1, rest_count1 = 0,0
                    sleep_time2, sleep_count2 = 0,0
                    rest_time2, rest_count2 = 0,0
                    sleep_time3, sleep_count3 = 0,0
                    rest_time3, rest_count3 = 0,0
                    one_two = 0
                    sum1 = 0
                    sum2 = 0
                    sum3 = 0
                    one_three = 0
                    two_three = 0
                    mm = 0

                else:
                    # 处理三只老鼠的帧数据
                    if all_coordinates_3m:
                        mouse1, mouse2, mouse3, sum1_1, sum2_1, sum3, one_two_1, one_three, two_three, sleep_1, rest_1 = behavior(
                            target_contents, all_coordinates_3m)
                        total_mouse1.extend(mouse1)
                        total_mouse2.extend(mouse2)
                        total_mouse3.extend(mouse3)
                    # 提取三只老鼠的行为数据
                    sleep_time1, sleep_count1 = sleep_1[0], sleep_1[3]
                    rest_time1, rest_count1 = rest_1[0], rest_1[3]
                    sleep_time2, sleep_count2 = sleep_1[1], sleep_1[4]
                    rest_time2, rest_count2 = rest_1[1], rest_1[4]
                    sleep_time3, sleep_count3 = sleep_1[2], sleep_1[5]
                    rest_time3, rest_count3 = rest_1[2], rest_1[5]
                    one_two = one_two_1
                    sum1 = sum1_1
                    sum2 = sum2_1
                    mm = mouse1.count('juji')
                print("3mouse done")

                if len(all_coordinates_2m) == 0:
                    pass
                else:
                    # 处理两只老鼠的帧数据
                    if all_coordinates_2m:
                        mouse1, mouse2, sum1_2, sum2_2, one_two_2, sleep_2, rest_2 = behavior2mous(target_contents,
                                                                                                     all_coordinates_2m)
                        total_mouse1.extend(mouse1)
                        total_mouse2.extend(mouse2)

                    # 提取两只老鼠的行为数据并将与前面的数据相加
                    sleep_time1, sleep_count1 = sleep_time1 + sleep_2[0], sleep_count1 + sleep_2[2]
                    rest_time1, rest_count1 = rest_time1 + rest_2[0], rest_count1 + rest_2[2]
                    sleep_time2, sleep_count2 = sleep_time2 + sleep_2[1], sleep_count2 + sleep_2[3]
                    rest_time2, rest_count2 = rest_time2 + rest_2[1], rest_count2 + rest_2[3]
                    one_two = one_two + one_two_2
                    sum1 = sum1 + sum1_2
                    sum2 = sum2 + sum2_2
                print("2mouse done")

                if len(all_coordinates_1m) == 0:
                    pass
                else:
                    # 处理一只老鼠的帧数据
                    if all_coordinates_1m:
                        mouse1, sum1_3, sleep_3, rest_3 = behavior1mous(target_contents, all_coordinates_1m)
                        total_mouse1.extend(mouse1)
                    print("1mouse done")
                    # 提取一只老鼠的行为数据并将与前面的数据相加
                    sleep_time1, sleep_count1 = sleep_time1 + sleep_3[0], sleep_count1 + sleep_3[1]
                    rest_time1, rest_count1 = rest_time1 + rest_3[0], rest_count1 + rest_3[1]
                    sum1 = sum1 + sum1_3

                mouse1_counts = {
                    'Eating': total_mouse1.count('Eating'),
                    'Drinking': total_mouse1.count('Drinking'),
                    'Climbing': total_mouse1.count('Climbing'),
                    'Curving': total_mouse1.count('Curving'),
                    'Elongating': total_mouse1.count('Elongating'),
                    'Shrinking': total_mouse1.count('Shrinking'),
                    'Grooming': total_mouse1.count('Grooming'),
                    'Standing': total_mouse1.count('Standing'),
                    'Stopping': total_mouse1.count('Stopping'),
                    'Walking': total_mouse1.count('Walking')  # ,
                    # 'juji': final_mouse1.count('juji')
                }
                write_mouse_data(ws, 2, mouse1_counts, sum1, rest_time1, rest_count1, sleep_time1, sleep_count1)
                ##这段用write_mouse_data替代
                row_start = 2  # 标题行之后的第一行
                # for index, (action, count) in enumerate(mouse1_counts.items(), start=2):
                #     ws.cell(row=row_start, column=index, value=count)

                # ws.cell(row=row_start, column=12, value=sum1)  # 行走次数
                # ws.cell(row=row_start, column=13, value=rest_time1)  # 休息时间
                # ws.cell(row=row_start, column=14, value=rest_count1)  # 休息次数
                # ws.cell(row=row_start, column=15, value=sleep_time1)  # 睡觉时间
                # ws.cell(row=row_start, column=16, value=sleep_count1)  # 睡觉次数
                ws.cell(row=row_start, column=17, value=one_two)  # mouse12
                ws.cell(row=row_start, column=18, value=one_three)  # mouse13
                ws.cell(row=row_start, column=19, value=two_three)  # mouse23

                ws.cell(row=row_start, column=20, value=mm)  # 三只小鼠聚集

                mouse2_counts = {
                    'Eating': total_mouse2.count('Eating'),
                    'Drinking': total_mouse2.count('Drinking'),
                    'Climbing': total_mouse2.count('Climbing'),
                    'Curving': total_mouse2.count('Curving'),
                    'Elongating': total_mouse2.count('Elongating'),
                    'Shrinking': total_mouse2.count('Shrinking'),
                    'Grooming': total_mouse2.count('Grooming'),
                    'Standing': total_mouse2.count('Standing'),
                    'Stopping': total_mouse2.count('Stopping'),
                    'Walking': total_mouse2.count('Walking')
                }

                # row_start = 3  # 标题行之后的第二行
                write_mouse_data(ws, 3, mouse2_counts, sum2, rest_time2, rest_count2, sleep_time2, sleep_count2)

                mouse3_counts = {
                    'Eating': total_mouse3.count('Eating'),
                    'Drinking': total_mouse3.count('Drinking'),
                    'Climbing': total_mouse3.count('Climbing'),
                    'Curving': total_mouse3.count('Curving'),
                    'Elongating': total_mouse3.count('Elongating'),
                    'Shrinking': total_mouse3.count('Shrinking'),
                    'Grooming': total_mouse3.count('Grooming'),
                    'Standing': total_mouse3.count('Standing'),
                    'Stopping': total_mouse3.count('Stopping'),
                    'Walking': total_mouse3.count('Walking')
                }

                # row_start = 4  # 标题行之后的第三行
                write_mouse_data(ws, 4, mouse3_counts, sum3, rest_time3, rest_count3, sleep_time3, sleep_count3)
                print("write done")
                # 保存 Excel 文件
                wb.save(save_path)

            print('完成')
    except FileNotFoundError:
        print("文件夹未找到，请检查路径。")
    except Exception as e:
        print(f"发生错误: {e}")


def write_mouse_data(ws, row_start, mouse_counts, sum_value, rest_time, rest_count, sleep_time, sleep_count):
    for index, (action, count) in enumerate(mouse_counts.items(), start=2):
        ws.cell(row=row_start, column=index, value=count)
    ws.cell(row=row_start, column=12, value=sum_value)  # 行走次数
    ws.cell(row=row_start, column=13, value=rest_time)  # 休息时间
    ws.cell(row=row_start, column=14, value=rest_count)  # 休息次数
    ws.cell(row=row_start, column=15, value=sleep_time)  # 睡觉时间
    ws.cell(row=row_start, column=16, value=sleep_count)  # 睡觉次数


def behavior(target_contents, all_coordinates):
    # target_contents点：饮水口、食物槽、饲养区
    mouse_one = []  # 创建一个空列表
    mouse_two = []
    mouse_three = []
    sleep_time1 = 0
    sleep_count1 = 0
    rest_time1 = 0
    rest_count1 = 0
    sleep_time2 = 0
    sleep_count2 = 0
    rest_time2 = 0
    rest_count2 = 0
    sleep_time3 = 0
    sleep_count3 = 0
    rest_time3 = 0
    rest_count3 = 0
    drink_point = target_contents[0]
    eat_point = target_contents[1:5]
    siyang_point = target_contents[5:9]
    max1 = 0
    max2 = 0
    max3 = 0
    sum1 = 0
    sum2 = 0
    sum3 = 0
    one_two = 0
    one_three = 0
    two_three = 0

    j = 0
    jj = 0
    jjj = 0
    for i in range(0, len(all_coordinates), 3):
        all_coordinate = all_coordinates[i:i + 3]
        if i // 3 + 1 == 1:
            pass
        else:
            # 第一只老鼠
            nose1 = all_coordinate[0][6:8]
            bozi1 = all_coordinate[0][12:14]
            center1 = all_coordinate[0][14:16]
            tail1 = all_coordinate[0][16:18]
            Nose_bozi_distance1 = calculate_distance(nose1, bozi1)
            Bozi_center_distance1 = calculate_distance(bozi1, center1)
            Center_tail_dinstance1 = calculate_distance(center1, tail1)
            Body_length1 = float(Nose_bozi_distance1 + Bozi_center_distance1 + Center_tail_dinstance1)
            if Body_length1 > max1:
                max1 = Body_length1

            # 第二只老鼠
            nose2 = all_coordinate[1][6:8]
            bozi2 = all_coordinate[1][12:14]
            center2 = all_coordinate[1][14:16]
            tail2 = all_coordinate[1][16:18]
            Nose_bozi_distance2 = calculate_distance(nose2, bozi2)
            Bozi_center_distance2 = calculate_distance(bozi2, center2)
            Center_tail_dinstance2 = calculate_distance(center2, tail2)
            Body_length2 = float(Nose_bozi_distance2 + Bozi_center_distance2 + Center_tail_dinstance2)
            if Body_length2 > max2:
                max2 = Body_length2

            # 第三只老鼠
            nose3 = all_coordinate[2][6:8]
            bozi3 = all_coordinate[2][12:14]
            center3 = all_coordinate[2][14:16]
            tail3 = all_coordinate[2][16:18]
            Nose_bozi_distance3 = calculate_distance(nose3, bozi3)
            Bozi_center_distance3 = calculate_distance(bozi3, center3)
            Center_tail_dinstance3 = calculate_distance(center3, tail3)
            Body_length3 = float(Nose_bozi_distance3 + Bozi_center_distance3 + Center_tail_dinstance3)
            if Body_length3 > max3:
                max3 = Body_length3

    for i in range(0, len(all_coordinates), 3):
        # 获取当前组的三个坐标
        # id为3其实是第一只老鼠
        all_coordinate = all_coordinates[i:i + 3]
        if i // 3 + 1 == 1:
            # speed = 0
            pass

        else:
            # 第一只老鼠
            kuang1 = all_coordinate[0][2:6]
            kuang_later1 = all_coordinates[i - 3][2:6]
            kuang_center_later1 = find_center(kuang_later1)
            kuang_center1 = find_center(kuang1)
            distance1 = calculate_distance(kuang_center_later1, kuang_center1)
            speed1 = distance1 / float(0.05)
            sum1 = sum1 + distance1

            # 第二只老鼠
            kuang2 = all_coordinate[1][2:6]
            kuang_later2 = all_coordinates[i - 2][2:6]
            kuang_center_later2 = find_center(kuang_later2)
            kuang_center2 = find_center(kuang2)
            distance2 = calculate_distance(kuang_center_later2, kuang_center2)
            speed2 = distance2 / float(0.05)
            sum2 = sum2 + distance2

            # 第三只老鼠
            kuang3 = all_coordinate[2][2:6]
            kuang_later3 = all_coordinates[i - 1][2:6]
            kuang_center_later3 = find_center(kuang_later3)
            kuang_center3 = find_center(kuang3)
            distance3 = calculate_distance(kuang_center_later3, kuang_center3)
            speed3 = distance3 / float(0.05)
            sum3 = sum3 + distance3

            # 行为分析
            behavior1 = behavior_fenxi(all_coordinate[0], all_coordinate[1], all_coordinate[2], eat_point, siyang_point,
                                     drink_point, speed1, i // 3 + 1, max1)
            mouse_one.append(behavior1)
            behavior2 = behavior_fenxi(all_coordinate[1], all_coordinate[0], all_coordinate[2], eat_point, siyang_point,
                                     drink_point, speed2, i // 3 + 1, max2)
            mouse_two.append(behavior2)
            behavior3 = behavior_fenxi(all_coordinate[2], all_coordinate[0], all_coordinate[1], eat_point, siyang_point,
                                     drink_point, speed3, i // 3 + 1, max3)
            mouse_three.append(behavior3)

            if IoUofBox(all_coordinate[0], all_coordinate[1]):
                one_two += 1

            if IoUofBox(all_coordinate[0], all_coordinate[2]):
                one_three += 1

            if IoUofBox(all_coordinate[1], all_coordinate[2]):
                two_three += 1

            # print(mouse_one, mouse_two, mouse_three)
    # 睡觉，休息
    for i in range(0, len(all_coordinates), 3):
        all_coordinate = all_coordinates[i:i + 3]
        if i // 3 + 1 == 1:
            pass
        elif i // 3 + 1 == len(all_coordinates) // 3:
            pass

        else:
            # mouse1
            kuang1 = all_coordinate[0][2:6]
            kuang_later1 = all_coordinates[i - 3][2:6]
            kuang_framer1 = all_coordinates[i + 3][2:6]
            kuang_center_later1 = find_center(kuang_later1)
            kuang_center1 = find_center(kuang1)
            kuang_center_framer1 = find_center(kuang_framer1)
            one1 = calculate_distance(kuang_center_later1, kuang_center1)
            two1 = calculate_distance(kuang_center_framer1, kuang_center1)

            if one1 < 10 and two1 < 10:
                j += 1

            else:
                if 400 <= j < 800:
                    rest_time1 = j + rest_time1
                    rest_count1 += 1
                elif j >= 800:
                    sleep_time1 = j + sleep_time1
                    sleep_count1 += 1
                else:
                    j = 0
                j = 0
            # print("i:",i)
            # print("j:",j)
            # mouse2
            kuang2 = all_coordinate[1][2:6]
            kuang_later2 = all_coordinates[i - 2][2:6]
            kuang_framer2 = all_coordinates[i + 4][2:6]
            kuang_center_later2 = find_center(kuang_later2)
            kuang_center2 = find_center(kuang2)
            kuang_center_framer2 = find_center(kuang_framer2)
            one2 = calculate_distance(kuang_center_later2, kuang_center2)
            two2 = calculate_distance(kuang_center_framer2, kuang_center2)

            if one2 < 10 and two2 < 10:
                jj += 1

            else:
                if 400 <= jj < 800:
                    rest_time2 = jj + rest_time2
                    rest_count2 += 1
                elif jj >= 800:
                    sleep_time2 = jj + sleep_time2
                    sleep_count2 += 1
                else:
                    jj = 0
                jj = 0

            # mouse3
            kuang3 = all_coordinate[2][2:6]
            kuang_later3 = all_coordinates[i - 1][2:6]
            kuang_framer3 = all_coordinates[i + 5][2:6]
            kuang_center_later3 = find_center(kuang_later3)
            kuang_center3 = find_center(kuang3)
            kuang_center_framer3 = find_center(kuang_framer3)
            one3 = calculate_distance(kuang_center_later3, kuang_center3)
            two3 = calculate_distance(kuang_center_framer3, kuang_center3)

            if one3 < 10 and two3 < 10:
                jjj += 1

            else:
                if 400 <= jjj < 800:
                    rest_time3 = jjj + rest_time3
                    rest_count3 += 1
                elif jjj >= 800:
                    sleep_time3 = jjj + sleep_time3
                    sleep_count3 += 1
                else:
                    jjj = 0
                jjj = 0
    sleep = [sleep_time1, sleep_time2, sleep_time3, sleep_count1, sleep_count2, sleep_count3]
    rest = [rest_time1, rest_time2, rest_time3, rest_count1, rest_count2, rest_count3]
    return mouse_one, mouse_two, mouse_three, sum1, sum2, sum3, one_two, one_three, two_three, sleep, rest


def behavior2mous(target_contents, all_coordinates):
    # target_contents点：饮水口、食物槽、饲养区
    mouse_one = []  # 创建一个空列表
    mouse_two = []
    sleep_time1 = 0
    sleep_count1 = 0
    rest_time1 = 0
    rest_count1 = 0
    sleep_time2 = 0
    sleep_count2 = 0
    rest_time2 = 0
    rest_count2 = 0
    drink_point = target_contents[0]
    eat_point = target_contents[1:5]
    siyang_point = target_contents[5:9]
    max1 = 0
    max2 = 0
    sum1 = 0
    sum2 = 0
    one_two = 0

    j = 0
    jj = 0

    for i in range(0, len(all_coordinates), 2):
        all_coordinate = all_coordinates[i:i + 2]
        if i in [0, 1]:
            pass
        else:
            # 第一只老鼠
            nose1 = all_coordinate[0][6:8]
            bozi1 = all_coordinate[0][12:14]
            center1 = all_coordinate[0][14:16]
            tail1 = all_coordinate[0][16:18]
            Nose_bozi_distance1 = calculate_distance(nose1, bozi1)
            Bozi_center_distance1 = calculate_distance(bozi1, center1)
            Center_tail_dinstance1 = calculate_distance(center1, tail1)
            Body_length1 = float(Nose_bozi_distance1 + Bozi_center_distance1 + Center_tail_dinstance1)
            if Body_length1 > max1:
                max1 = Body_length1

            # 第二只老鼠
            nose2 = all_coordinate[1][6:8]
            bozi2 = all_coordinate[1][12:14]
            center2 = all_coordinate[1][14:16]
            tail2 = all_coordinate[1][16:18]
            Nose_bozi_distance2 = calculate_distance(nose2, bozi2)
            Bozi_center_distance2 = calculate_distance(bozi2, center2)
            Center_tail_dinstance2 = calculate_distance(center2, tail2)
            Body_length2 = float(Nose_bozi_distance2 + Bozi_center_distance2 + Center_tail_dinstance2)
            if Body_length2 > max2:
                max2 = Body_length2

    for i in range(0, len(all_coordinates), 2):
        # 获取当前组的三个坐标
        # id为3其实是第一只老鼠
        all_coordinate = all_coordinates[i:i + 2]
        if i in [0, 1]:
            speed = 0

        else:
            # 第一只老鼠
            kuang1 = all_coordinate[0][2:6]
            kuang_later1 = all_coordinates[i - 2][2:6]
            kuang_center_later1 = find_center(kuang_later1)
            kuang_center1 = find_center(kuang1)
            distance1 = calculate_distance(kuang_center_later1, kuang_center1)
            speed1 = distance1 / float(0.05)
            sum1 = sum1 + distance1

            # 第二只老鼠
            kuang2 = all_coordinate[1][2:6]
            kuang_later2 = all_coordinates[i - 1][2:6]
            kuang_center_later2 = find_center(kuang_later2)
            kuang_center2 = find_center(kuang2)
            distance2 = calculate_distance(kuang_center_later2, kuang_center2)
            speed2 = distance2 / float(0.05)
            sum2 = sum2 + distance2

            # 行为分析
            behavior1 = behavior_fenxi1(all_coordinate[0], eat_point, siyang_point, drink_point, speed1, i // 2 + 1, max1)
            mouse_one.append(behavior1)
            behavior2 = behavior_fenxi1(all_coordinate[1], eat_point, siyang_point, drink_point, speed2, i // 2 + 1, max2)
            mouse_two.append(behavior2)

            if IoUofBox(all_coordinate[0], all_coordinate[1]):
                one_two += 1

            # print(mouse_one, mouse_two, mouse_three)
    # 睡觉，休息
    for i in range(0, len(all_coordinates), 2):
        all_coordinate = all_coordinates[i:i + 2]
        if i in [0, 1]:
            pass
        elif i // 2 + 1 == len(all_coordinates) // 2:
            pass

        else:
            # mouse1
            kuang1 = all_coordinate[0][2:6]
            kuang_later1 = all_coordinates[i - 2][2:6]
            kuang_framer1 = all_coordinates[i + 2][2:6]
            kuang_center_later1 = find_center(kuang_later1)
            kuang_center1 = find_center(kuang1)
            kuang_center_framer1 = find_center(kuang_framer1)
            one1 = calculate_distance(kuang_center_later1, kuang_center1)
            two1 = calculate_distance(kuang_center_framer1, kuang_center1)

            if one1 < 10 and two1 < 10:
                j += 1

            else:
                if 400 <= j < 800:
                    rest_time1 = j + rest_time1
                    rest_count1 += 1
                elif j >= 800:
                    sleep_time1 = j + sleep_time1
                    sleep_count1 += 1
                else:
                    j = 0
                j = 0

            # mouse2
            kuang2 = all_coordinate[1][2:6]
            kuang_later2 = all_coordinates[i - 1][2:6]
            kuang_framer2 = all_coordinates[i + 3][2:6]
            kuang_center_later2 = find_center(kuang_later2)
            kuang_center2 = find_center(kuang2)
            kuang_center_framer2 = find_center(kuang_framer2)
            one2 = calculate_distance(kuang_center_later2, kuang_center2)
            two2 = calculate_distance(kuang_center_framer2, kuang_center2)

            if one2 < 10 and two2 < 10:
                jj += 1

            else:
                if 400 <= jj < 800:
                    rest_time2 = jj + rest_time2
                    rest_count2 += 1
                elif jj >= 800:
                    sleep_time2 = jj + sleep_time2
                    sleep_count2 += 1
                else:
                    jj = 0
                jj = 0

    sleep = [sleep_time1, sleep_time2, sleep_count1, sleep_count2]
    rest = [rest_time1, rest_time2, rest_count1, rest_count2]
    return mouse_one, mouse_two, sum1, sum2, one_two, sleep, rest


def behavior1mous(target_contents, all_coordinates):
    # target_contents点：饮水口、食物槽、饲养区
    mouse_one = []  # 创建一个空列表
    sleep_time1 = 0
    sleep_count1 = 0
    rest_time1 = 0
    rest_count1 = 0
    drink_point = target_contents[0]
    eat_point = target_contents[1:5]
    siyang_point = target_contents[5:9]
    max1 = 0
    sum1 = 0

    j = 0
    for i in range(0, len(all_coordinates), 1):
        all_coordinate = all_coordinates[i]
        if i + 1 == 1:
            pass
        else:
            # 第一只老鼠
            nose1 = all_coordinate[6:8]
            bozi1 = all_coordinate[12:14]
            center1 = all_coordinate[14:16]
            tail1 = all_coordinate[16:18]
            Nose_bozi_distance1 = calculate_distance(nose1, bozi1)
            Bozi_center_distance1 = calculate_distance(bozi1, center1)
            Center_tail_dinstance1 = calculate_distance(center1, tail1)
            Body_length1 = float(Nose_bozi_distance1 + Bozi_center_distance1 + Center_tail_dinstance1)
            if Body_length1 > max1:
                max1 = Body_length1

    for i in range(0, len(all_coordinates), 1):
        # 获取当前组的三个坐标
        # id为3其实是第一只老鼠
        all_coordinate = all_coordinates[i]
        if i + 1 == 1:
            speed = 0

        else:
            # 第一只老鼠
            kuang1 = all_coordinate[2:6]
            kuang_later1 = all_coordinates[i - 1][2:6]
            kuang_center_later1 = find_center(kuang_later1)
            kuang_center1 = find_center(kuang1)
            distance1 = calculate_distance(kuang_center_later1, kuang_center1)
            speed1 = distance1 / float(0.05)
            sum1 = sum1 + distance1

            # 行为分析
            behavior1 = behavior_fenxi1(all_coordinate, eat_point, siyang_point, drink_point, speed1, i + 1, max1)
            mouse_one.append(behavior1)

            # print(mouse_one, mouse_two, mouse_three)
    # 睡觉，休息
    for i in range(0, len(all_coordinates), 1):
        all_coordinate = all_coordinates[i]
        if i + 1 == 1:
            pass
        elif i + 1 == len(all_coordinates):
            pass

        else:
            # mouse1
            kuang1 = all_coordinate[2:6]
            kuang_later1 = all_coordinates[i - 1][2:6]
            kuang_framer1 = all_coordinates[i + 1][2:6]
            kuang_center_later1 = find_center(kuang_later1)
            kuang_center1 = find_center(kuang1)
            kuang_center_framer1 = find_center(kuang_framer1)
            one1 = calculate_distance(kuang_center_later1, kuang_center1)
            two1 = calculate_distance(kuang_center_framer1, kuang_center1)

            if one1 < 10 and two1 < 10:
                j += 1

            else:
                if 400 <= j < 800:
                    rest_time1 = j + rest_time1
                    rest_count1 += 1
                elif j >= 800:
                    sleep_time1 = j + sleep_time1
                    sleep_count1 += 1
                else:
                    j = 0
                j = 0

    sleep = [sleep_time1, sleep_count1]
    rest = [rest_time1, rest_count1]
    return mouse_one, sum1, sleep, rest


def behavior_fenxi(all_coordinate, pose2, pose3, eat_point, siyang_point, drink_point, speed, i, max):
    # 夹角阈值
    angle_threshold = 30
    angle_threshold1 = 80

    # 关键点参数
    nose = all_coordinate[6:8]
    bozi = all_coordinate[12:14]
    center = all_coordinate[14:16]
    tail = all_coordinate[16:18]
    # 头部偏角
    angle = jiaodu(nose, bozi, center)
    # 与饮水口距离
    drink_distance = calculate_distance(nose, drink_point)
    # 身体长度
    Nose_bozi_distance = calculate_distance(nose, bozi)
    Bozi_center_distance = calculate_distance(bozi, center)
    Center_tail_dinstance = calculate_distance(center, tail)
    Body_length = float(Nose_bozi_distance + Bozi_center_distance + Center_tail_dinstance)
    # 头尾长度
    Nose_tail_distance = calculate_distance(nose, tail)
    # 判断伸长阈值
    # Nose_bozi_distance_baifenlv = float(Nose_bozi_distance / Body_length)
    Body_length_baifenlv = float(Nose_tail_distance / Body_length)

    # bozi_tail_baifenlv = float(Nose_bozi_distance / Nose_tail_distance)

    # 判断蜷缩
    d = float(Body_length / max)
    # 伸直判断
    Body_length_threshold = 0.95
    # 蜷缩判断
    Curving_threshold = 0.6

    # juxing_distance_file_path = r'D:\software\behavior.txt'
    #
    # with open(juxing_distance_file_path, "a") as f:
    #     f.write(f'zhenlv={i}，d={d}，max_Body_length3={max}，Body_length={Body_length}' + "\n")

    # 判断是不是在饲养区
    if is_point_in_polygon(center, siyang_point):

        if is_point_in_polygon(nose, eat_point):
            # 判断有没有在进食区
            if is_point_in_polygon(center, eat_point):

                return 'Climbing'
            else:
                return 'Eating'
        # 判断有没有在饮水区
        elif drink_distance < int(20):
            return 'Drinking'

        else:
            # 判断是不是三只聚集
            if juji(all_coordinate, pose2, pose3):

                return 'juji'

            else:
                # 判断是不是行走
                if int(speed) < int(200):
                    # 判断是不是伸长
                    if Body_length_baifenlv > Body_length_threshold:

                        return 'Elongating'
                    # 判断是不是弯曲
                    elif d <= Curving_threshold:

                        return 'Curving'
                    # 判断是不是蜷缩
                    elif angle_threshold < angle < angle_threshold1:

                        return 'Shrinking'
                    # 判断是不是梳理
                    elif angle >= angle_threshold1:

                        return 'Grooming'

                    else:

                        return 'Stopping'

                else:

                    return 'Walking'

    else:
        # 判断是不是攀爬
        if is_point_in_polygon(center, eat_point):

            return 'Climbing'

        else:

            return 'Standing'


def behavior_fenxi1(all_coordinate, eat_point, siyang_point, drink_point, speed, i, max):
    # 夹角阈值
    angle_threshold = 30
    angle_threshold1 = 80

    # 关键点参数
    nose = all_coordinate[6:8]
    bozi = all_coordinate[12:14]
    center = all_coordinate[14:16]
    tail = all_coordinate[16:18]
    # 头部偏角
    angle = jiaodu(nose, bozi, center)
    # 与饮水口距离
    drink_distance = calculate_distance(nose, drink_point)
    # 身体长度
    Nose_bozi_distance = calculate_distance(nose, bozi)
    Bozi_center_distance = calculate_distance(bozi, center)
    Center_tail_dinstance = calculate_distance(center, tail)
    Body_length = float(Nose_bozi_distance + Bozi_center_distance + Center_tail_dinstance)
    # 头尾长度
    Nose_tail_distance = calculate_distance(nose, tail)
    # 判断伸长阈值
    # Nose_bozi_distance_baifenlv = float(Nose_bozi_distance / Body_length)
    Body_length_baifenlv = float(Nose_tail_distance / Body_length)

    # bozi_tail_baifenlv = float(Nose_bozi_distance / Nose_tail_distance)

    # 判断蜷缩
    d = float(Body_length / max)
    # 伸直判断
    Body_length_threshold = 0.95
    # 蜷缩判断
    Curving_threshold = 0.6

    # juxing_distance_file_path = r'D:\software\behavior.txt'
    #
    # with open(juxing_distance_file_path, "a") as f:
    #     f.write(f'zhenlv={i}，d={d}，max_Body_length3={max}，Body_length={Body_length}' + "\n")

    # 判断是不是在饲养区
    if is_point_in_polygon(center, siyang_point):

        if is_point_in_polygon(nose, eat_point):
            # 判断有没有在进食区
            if is_point_in_polygon(center, eat_point):

                return 'Climbing'
            else:
                return 'Eating'
        # 判断有没有在饮水区
        elif drink_distance < int(20):
            return 'Drinking'

        else:
            # 判断是不是行走
            if int(speed) < int(200):
                # 判断是不是伸长
                if Body_length_baifenlv > Body_length_threshold:

                    return 'Elongating'
                # 判断是不是弯曲
                elif d <= Curving_threshold:

                    return 'Curving'
                # 判断是不是蜷缩
                elif angle_threshold < angle < angle_threshold1:

                    return 'Shrinking'
                # 判断是不是梳理
                elif angle >= angle_threshold1:

                    return 'Grooming'

                else:

                    return 'Stopping'

            else:

                return 'Walking'

    else:
        # 判断是不是攀爬
        if is_point_in_polygon(center, eat_point):

            return 'Climbing'

        else:

            return 'Standing'


# 判断三只老鼠是不是聚集
def juji(pose1, pose2, pose3):
    A = IoUofBox(pose1, pose2)
    B = IoUofBox(pose1, pose3)
    C = IoUofBox(pose3, pose2)
    D = A + B + C
    # print('111', A, B, C, D)
    if D > 1:

        return 1

    else:
        return 0


def calculate_distance(point1, point2):
    """计算两个点之间的距离"""
    return int(math.sqrt((point2[0] - point1[0]) ** 2 + (point2[1] - point1[1]) ** 2))


def IoUofBox(pos1, pos2):
    # 两个矩形区域坐标提取
    xa1 = pos1[2]
    ya1 = pos1[3]
    xa2 = pos1[4]
    ya2 = pos1[5]
    xb1 = pos2[2]
    yb1 = pos2[3]
    xb2 = pos2[4]
    yb2 = pos2[5]
    # 交集区域坐标
    xI1 = max(xa1, xb1)
    yI1 = max(ya1, yb1)
    xI2 = min(xa2, xb2)
    yI2 = min(ya2, yb2)
    # 交集区域面积
    SI = max((xI2 - xI1), 0) * max((yI2 - yI1), 0)
    # 并集区域面积
    SA = (xa2 - xa1) * (ya2 - ya1)
    SB = (xb2 - xb1) * (yb2 - yb1)
    SU = SA + SB - SI
    if SU == 0:
        return 0

    IoU = SI / SU

    if IoU > 0.2:
        return 1
    else:
        return 0
    # return IoU


# 判断是不是站立
def Stand_against_wall(nose_point, siyang):
    point_x, point_y = nose_point[0][0], nose_point[0][1]
    x1, x3 = 600, 1300
    y1, y3 = 200, 920
    if x1 <= point_x <= x3 and y1 <= point_y <= y3:
        return 0
    else:
        return 1


# 判断点是不是在食物槽里面
def is_point_in_polygon(point, eat_point):
    x, y = point
    n = len(eat_point)
    inside = False

    p1x, p1y = eat_point[0]
    for i in range(n + 1):
        p2x, p2y = eat_point[i % n]
        if y > min(p1y, p2y):
            if y <= max(p1y, p2y):
                if x <= max(p1x, p2x):
                    if p1y != p2y:
                        xinters = (y - p1y) * (p2x - p1x) / (p2y - p1y) + p1x
                    if p1x == p2x or x <= xinters:
                        inside = not inside
        p1x, p1y = p2x, p2y
    return inside


# 计算角度
def jiaodu(pointa, pointb, pointc):
    # 鼻子， 脖子， 身体中心
    try:
        # 计算向量 AB 和 BC
        vector_ab = [pointb[0] - pointa[0], pointb[1] - pointa[1]]
        vector_bc = [pointc[0] - pointb[0], pointc[1] - pointb[1]]

        # 计算向量 AB 和 BC 的长度
        length_ab = math.sqrt(vector_ab[0] ** 2 + vector_ab[1] ** 2)
        length_bc = math.sqrt(vector_bc[0] ** 2 + vector_bc[1] ** 2)

        # 防止除以零
        if length_ab == 0 or length_bc == 0:
            return 0  # 或者返回某个特定的值，表示无法计算角度

        # 计算向量 AB 和 BC 的点积
        dot_product = vector_ab[0] * vector_bc[0] + vector_ab[1] * vector_bc[1]

        # 计算夹角的余弦值
        cos_angle = dot_product / (length_ab * length_bc)

        # 限制 cos_angle 的范围在 [-1, 1] 之间，以避免 acos 的错误
        cos_angle = max(-1, min(1, cos_angle))

        # 使用反余弦函数计算夹角（弧度）
        angle_rad = math.acos(cos_angle)
        angle_deg = math.degrees(angle_rad)
        return angle_deg

    except ValueError:
        return None  # 返回 None 表示计算出错
    except TypeError:
        return None  # 返回 None 表示输入类型错误


# 框的中心点
def find_center(kuang):
    if len(kuang) != 4:
        raise ValueError("必须提供四个点")

    # 解包四个点
    x1, y1, x2, y2 = kuang[0:4]

    # 计算中心点坐标
    center_x = int((x1 + x2) / 2)
    center_y = int((y1 + y2) / 2)

    return (center_x, center_y)

if __name__ == '__main__':
    # 指定文件夹路径
    folder_path = r'/home/shanli/video_analysis/video_result_txt/male_20230825'  # 关键点文件夹
    target_folder = r'/home/shanli/video_analysis/plot_label/position.txt/male_20230825'  # 环境关键点文件夹
    # 行为数据保存文件夹=
    save_folder = r'/home/shanli/video_analysis/behavior_result/male_20230825'


    txt(folder_path, target_folder, save_folder)

